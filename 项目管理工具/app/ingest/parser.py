"""单文件解析器: 打开 xlsx, 解析 4 个区块 → Project + QcIssue[]。

键值解析: 逐行扫描所有列找标签; 值 = 其右侧起、到本行下一个标签单元格之前区间内第一个非空值。
区块锚点: 仅 A 列扫描。
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any

import openpyxl

from app.common.models import Project, Member, Task, QcIssue
from app.common.dates import parse_and_fmt
from app.common.enums import (
    BIZ_ID_PATTERN,
    FILENAME_PATTERN,
    EXTERNAL_MEMBER_NAMES,
    MEMBER_NAME_SEPARATORS,
)
from app.ingest.template_map import (
    PRIMARY_SHEET,
    IGNORE_SHEETS,
    DATA_ONLY,
    SECTION_ANCHORS,
    ANCHOR_TEXTS,
    normalize_label,
    KEY_VALUE_FIELDS,
    MEMBER_HEADER_DETECT,
    MEMBER_COLUMN_MAP,
    TASK_HEADER_DETECT,
    TASK_COLUMN_MAP,
)


def _strip_cell(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(v: Any) -> str | None:
    return parse_and_fmt(v)


def _is_anchor_row(cell_text: str) -> bool:
    """判断 A 列单元格文本是否为区块锚点。"""
    return any(anchor in cell_text for anchor in ANCHOR_TEXTS)


def _match_key_value_label(normalized_label: str) -> dict | None:
    for key, field_def in KEY_VALUE_FIELDS.items():
        if key in normalized_label or normalized_label in key:
            return field_def
    return None


# ── 主入口 ──────────────────────────────────────────────

def parse(filename_meta: dict, filepath: Path) -> tuple[Project, list[QcIssue]]:
    issues: list[QcIssue] = []
    source = filename_meta["source_filename"]

    # 1. 打开工作簿
    try:
        wb = openpyxl.load_workbook(filepath, data_only=DATA_ONLY)
    except Exception:
        return (
            Project(biz_id="", snap_date=""),
            [QcIssue(
                source_filename=source, severity="error",
                issue_type="FILE_OPEN_FAIL", message=f"无法打开: {filepath}",
            )],
        )

    # 2. 定位 sheet
    sheet_names = wb.sheetnames
    sheet_name = PRIMARY_SHEET if PRIMARY_SHEET in sheet_names else sheet_names[0]
    ws = wb[sheet_name]
    if sheet_name != PRIMARY_SHEET and sheet_name not in IGNORE_SHEETS:
        issues.append(QcIssue(
            source_filename=source, severity="warning",
            issue_type="SHEET_MISSING", location="文件级",
            message=f"无「{PRIMARY_SHEET}」, 回退到「{sheet_name}」",
        ))

    # 3. 扫描 A 列找区块锚点
    anchors: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        cell_val = _strip_cell(ws.cell(row=row, column=1).value)
        if cell_val:
            for key, anchor_text in SECTION_ANCHORS.items():
                if anchor_text in cell_val:
                    anchors[key] = row
                    break

    # 4. 解析键值对(区块一/二) — 逐行扫描所有列
    project_data: dict[str, Any] = {}
    for section_key in ("basic", "contract"):
        start_row = anchors.get(section_key)
        if start_row is None:
            continue
        end_row = _find_next_anchor(start_row + 1, anchors)
        kv_issues = _parse_kv_rows(ws, start_row + 1, end_row, source, project_data)
        issues.extend(kv_issues)

    # 5. 解析人员分工(区块三)
    members: list[Member] = []
    if anchors.get("members"):
        members, member_issues = _parse_members(ws, anchors, source)
        issues.extend(member_issues)

    # 6. 解析阶段任务(区块四)
    tasks: list[Task] = []
    if anchors.get("tasks"):
        tasks, task_issues = _parse_tasks(ws, anchors, source)
        issues.extend(task_issues)

    # 7. 业务ID 质检
    biz_id = project_data.get("biz_id") or ""
    if not biz_id:
        issues.append(QcIssue(
            source_filename=source, severity="error",
            issue_type="BIZ_ID_MISSING", location="区块一/业务ID",
            message="表内业务ID为空",
        ))
    elif not BIZ_ID_PATTERN.match(biz_id):
        issues.append(QcIssue(
            source_filename=source, severity="warning",
            issue_type="BIZ_ID_FORMAT", location="区块一/业务ID",
            message=f"业务ID格式不符: {biz_id}",
        ))

    file_biz_id = filename_meta.get("biz_id_from_name", "")
    if file_biz_id and biz_id and file_biz_id != biz_id:
        issues.append(QcIssue(
            source_filename=source, severity="warning",
            issue_type="ID_FILENAME_MISMATCH", location="文件名 vs 单元格",
            message=f"文件名ID={file_biz_id}, 表内ID={biz_id}",
        ))
    if filename_meta.get("filename_format_issue"):
        issues.append(QcIssue(
            source_filename=source, severity="warning",
            issue_type="FILENAME_FORMAT", location="文件名",
            message=filename_meta["filename_format_issue"],
        ))

    # 8. 组装 Project
    proj = Project(
        biz_id=biz_id,
        snap_date=filename_meta.get("snap_date", ""),
        project_name=project_data.get("project_name"),
        biz_type=project_data.get("biz_type"),
        stage_status=project_data.get("stage_status"),
        customer=project_data.get("customer"),
        customer_industry=project_data.get("customer_industry"),
        description=project_data.get("description"),
        deliverables=project_data.get("deliverables"),
        setup_date=project_data.get("setup_date"),
        plan_deliver_date=project_data.get("plan_deliver_date"),
        actual_deliver_date=project_data.get("actual_deliver_date"),
        progress_pct=project_data.get("progress_pct"),
        risk_level=project_data.get("risk_level"),
        risk_desc=project_data.get("risk_desc"),
        current_issue=project_data.get("current_issue"),
        contract_status=project_data.get("contract_status"),
        contract_no=project_data.get("contract_no"),
        contract_amount=project_data.get("contract_amount"),
        invoiced_amount=project_data.get("invoiced_amount"),
        received_amount=project_data.get("received_amount"),
        gross_margin_pct=project_data.get("gross_margin_pct"),
        owner_name=project_data.get("owner_name"),
        pm_name=project_data.get("pm_name"),
        last_update_date=project_data.get("last_update_date"),
        form_status=project_data.get("form_status"),
        source_filename=source,
    )
    proj.members = members
    proj.tasks = tasks
    return proj, issues


# ── 新键值解析: 逐行扫描所有列 ──────────────────────────

def _parse_kv_rows(
    ws, start_row: int, end_row: int, source: str, project_data: dict
) -> list[QcIssue]:
    """在 [start_row, end_row) 内逐行扫描所有列找标签, 按新规则取右侧值。"""
    issues: list[QcIssue] = []

    for row in range(start_row, end_row):
        # 收集本行所有"候选标签单元": (col, raw_text)
        label_candidates: list[tuple[int, str]] = []
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=row, column=col).value
            txt = _strip_cell(raw)
            if not txt:
                continue
            label_candidates.append((col, txt))

        if not label_candidates:
            continue

        # 对每个候选, 尝试匹配 KEY_VALUE_FIELDS; 非标签的文本可能是区块锚点或普通文本, 跳过
        for i, (label_col, raw_text) in enumerate(label_candidates):
            normalized = normalize_label(raw_text)
            field_def = _match_key_value_label(normalized)
            if field_def is None:
                # 本单元格不是已知标签, 不用做标签处理。其右侧不用它作边界
                continue

            # 找"本行下一个标签"的列号(若有)
            next_label_col = None
            for j in range(i + 1, len(label_candidates)):
                ncol, ntext = label_candidates[j]
                nnorm = normalize_label(ntext)
                if _match_key_value_label(nnorm) is not None:
                    next_label_col = ncol
                    break

            # 在 (label_col+1, next_label_col) 区间内找第一个非空值
            value = None
            search_end = next_label_col if next_label_col else ws.max_column + 1
            for vcol in range(label_col + 1, search_end):
                v = ws.cell(row=row, column=vcol).value
                if v is not None and str(v).strip():
                    value = v
                    break

            # 类型转换
            col_name = field_def["column"]
            ftype = field_def["type"]
            if ftype == "date":
                converted = _parse_date(value)
                if value is not None and _strip_cell(value) and converted is None:
                    issues.append(QcIssue(
                        source_filename=source, severity="warning",
                        issue_type="BAD_DATE", location=f"键值区/{normalized}",
                        message=f"日期无法解析: {_strip_cell(value)}",
                    ))
                project_data[col_name] = converted
            elif ftype == "number":
                converted = _parse_number(value)
                if value is not None and _strip_cell(value) and converted is None:
                    issues.append(QcIssue(
                        source_filename=source, severity="warning",
                        issue_type="BAD_NUMBER", location=f"键值区/{normalized}",
                        message=f"数字无法解析: {_strip_cell(value)}",
                    ))
                project_data[col_name] = converted
            else:
                project_data[col_name] = _strip_cell(value)

            # 必填检查
            if field_def.get("required") and not project_data[col_name]:
                issues.append(QcIssue(
                    source_filename=source, severity="warning",
                    issue_type="REQUIRED_EMPTY", location=f"键值区/{normalized}",
                    message=f"必填字段为空: {normalized}",
                ))

    return issues


# ── 辅助/表解析 — 不变 ──────────────────────────────────

def _find_next_anchor(start_row: int, anchors: dict[str, int]) -> int:
    rows = [r for r in anchors.values() if r >= start_row]
    return min(rows) if rows else 99999


def _detect_header_row(ws, start_row: int, end_row: int, keywords: list[str]) -> int | None:
    for row in range(start_row, end_row):
        row_texts = {
            _strip_cell(ws.cell(row=row, column=c).value) or ""
            for c in range(1, ws.max_column + 1)
        }
        if all(any(kw in t for t in row_texts) for kw in keywords):
            return row
    return None


def _map_columns(ws, header_row: int, col_map: dict[str, str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        hdr = _strip_cell(ws.cell(row=header_row, column=col).value) or ""
        for keyword, field_name in col_map.items():
            if keyword in hdr:
                mapping[field_name] = col
                break
    return mapping


def _parse_members(ws, anchors: dict[str, int], source: str) -> tuple[list[Member], list[QcIssue]]:
    members: list[Member] = []
    issues: list[QcIssue] = []
    start = anchors["members"] + 1
    end = _find_next_anchor(start, anchors)

    header = _detect_header_row(ws, start, end, MEMBER_HEADER_DETECT)
    if header is None:
        return members, issues
    col_map = _map_columns(ws, header, MEMBER_COLUMN_MAP)

    for row in range(header + 1, end):
        if all(
            ws.cell(row=row, column=c).value is None
            or str(ws.cell(row=row, column=c).value).strip() == ""
            for c in range(1, ws.max_column + 1)
        ):
            break

        name = _strip_cell(ws.cell(row=row, column=col_map.get("name", 0)).value)

        is_ext = False
        if name is None or name.strip() in EXTERNAL_MEMBER_NAMES:
            is_ext = True
            if name and name.strip():
                issues.append(QcIssue(
                    source_filename=source, severity="warning",
                    issue_type="MEMBER_EXTERNAL", location=f"人员区行{row}",
                    message=f"人员姓名为{name}",
                ))

        if name and any(sep in name for sep in MEMBER_NAME_SEPARATORS):
            issues.append(QcIssue(
                source_filename=source, severity="warning",
                issue_type="MULTI_NAME_IN_MEMBER", location=f"人员区行{row}",
                message=f"人员姓名一格多名字: {name}",
            ))

        workload = _parse_number(ws.cell(row=row, column=col_map.get("workload_pct", 0)).value)
        if name and not is_ext and workload is None:
            issues.append(QcIssue(
                source_filename=source, severity="warning",
                issue_type="WORKLOAD_MISSING", location=f"人员区行{row}",
                message=f"{name} 缺少投入工作量",
            ))

        members.append(Member(
            row_idx=row,
            name=name,
            emp_id=_strip_cell(ws.cell(row=row, column=col_map.get("emp_id", 0)).value),
            role=_strip_cell(ws.cell(row=row, column=col_map.get("role", 0)).value),
            task_desc=_strip_cell(ws.cell(row=row, column=col_map.get("task_desc", 0)).value),
            join_start=_parse_date(ws.cell(row=row, column=col_map.get("join_start", 0)).value),
            join_end=_parse_date(ws.cell(row=row, column=col_map.get("join_end", 0)).value),
            workload_pct=workload,
            progress_note=_strip_cell(ws.cell(row=row, column=col_map.get("progress_note", 0)).value),
            eval=_strip_cell(ws.cell(row=row, column=col_map.get("eval", 0)).value),
            note=_strip_cell(ws.cell(row=row, column=col_map.get("note", 0)).value),
            is_external=is_ext,
        ))
    return members, issues


def _parse_tasks(ws, anchors: dict[str, int], source: str) -> tuple[list[Task], list[QcIssue]]:
    tasks: list[Task] = []
    issues: list[QcIssue] = []
    start = anchors["tasks"] + 1
    end = ws.max_row + 1

    header = _detect_header_row(ws, start, end, TASK_HEADER_DETECT)
    if header is None:
        return tasks, issues
    col_map = _map_columns(ws, header, TASK_COLUMN_MAP)

    for row in range(header + 1, end + 1):
        if all(
            ws.cell(row=row, column=c).value is None
            or str(ws.cell(row=row, column=c).value).strip() == ""
            for c in range(1, ws.max_column + 1)
        ):
            break

        task_name = _strip_cell(ws.cell(row=row, column=col_map.get("task_name", 0)).value)
        if not task_name:
            continue

        tasks.append(Task(
            row_idx=row,
            seq=_strip_cell(ws.cell(row=row, column=col_map.get("seq", 0)).value),
            task_name=task_name,
            owner=_strip_cell(ws.cell(row=row, column=col_map.get("owner", 0)).value),
            plan_start=_parse_date(ws.cell(row=row, column=col_map.get("plan_start", 0)).value),
            plan_end=_parse_date(ws.cell(row=row, column=col_map.get("plan_end", 0)).value),
            actual_start=_parse_date(ws.cell(row=row, column=col_map.get("actual_start", 0)).value),
            progress_pct=_parse_number(ws.cell(row=row, column=col_map.get("progress_pct", 0)).value),
            status=_strip_cell(ws.cell(row=row, column=col_map.get("status", 0)).value),
            note=_strip_cell(ws.cell(row=row, column=col_map.get("note", 0)).value),
        ))
    return tasks, issues
